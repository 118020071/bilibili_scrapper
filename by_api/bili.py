import requests, random, selenium, bs4, time, openpyxl, datetime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook

# Section: get detailed info via API

g_headers = [
    "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
    'Opera/9.25 (Windows NT 5.1; U; en)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
    'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9',
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "
]

def get_json(bvid, url):

    global g_headers

    url = url+bvid

    fakeua = { "User-Agent":g_headers[random.randint(0, len(g_headers)-1)] }

    html = requests.get(
        url=url,
        headers=fakeua
    )

    data = html.json()

    #return dict{}
    return data

def get_aid(bvid):

    url = "https://api.bilibili.com/x/web-interface/view?bvid="
    
    data = get_json(bvid, url)
    aid = data.get("data").get("aid")

    return aid

def get_videoInfo(bvid):
    
    url = "https://api.bilibili.com/x/web-interface/archive/stat?bvid="

    data = get_json(bvid, url)
    
    # form: aid,bvid,view,danmaku,reply,favorite,coin,share,like,now_rank,
    # his_rank,no_reprint,copyright,argue_msg,evaluation
    video_data = data.get("data")
    #copyright == 1: 自制, 2: 转载

    # print(video_data)
    return video_data

def get_tagInfo(bvid):
    
    url = "https://api.bilibili.com/x/web-interface/view/detail/tag?bvid="
    
    data = get_json(bvid, url)

    tag_data = data.get("data")

    whole_list = []
    tag_list = []

    for i in range(len(tag_data)):
        single_tag = tag_data[i]
        single_dict = {}
        single_dict["标签名称"] = single_tag.get("tag_name")
        single_dict["标签订阅数"] = single_tag.get("subscribed_count")
        single_dict["标签下视频数"] = single_tag.get("archive_count")
        single_dict["标签下精选视频数"] = single_tag.get("featured_count")

        whole_list.append(str(single_dict))
        tag_list.append(single_tag.get("tag_name"))

    # print(';'.join(whole_list), ','.join(tag_list))
    return ';'.join(whole_list), ','.join(tag_list)

def get_up_info(uuid):

    url = "https://api.bilibili.com/x/relation/stat?vmid="

    data = get_json(uuid, url)

    up_data = data.get("data")

    up_info = {}

    up_info["following"] = up_data.get("following")
    up_info["follower"] = up_data.get("follower")

    # print(up_info)
    return up_info

# End section

# Section: reform data

def deleteUnit(a):
    #例如有'万',则添加0
    if "万" in a:
        aTenThousand = a[0:a.find(".")]
        aLeftover = a[(a.find(".")+1):a.find("万")]
        a = int(aTenThousand)*10000 + int(aLeftover)*1000
    return str(a)

def getBVNum(a):
    #get BV号 from link url
    a = a.split("/")
    a = a[-1].split("?")
    a = a[0]

    return a

def getFullLink(a):
    #get full link from link url
    a = a.split("?")
    a = a[0]
    a = "https:"+a
    return a

def getUpUUID(a):
    #input up link
    back = a.split("/")[-1]
    font = back.split("?")[0]

    return font

def unixTimeDecoding(value):
    #convert time in unix to yy:mm:dd hh:mm:ss
    format = '%Y-%m-%d %H:%M:%S'
    value = time.localtime(value)
    dt = time.strftime(format, value)

    return dt

# End section

# Section: get video BV id list using selenium

def getVideoLinks(html):

    global videoList

    parseHTML = bs4.BeautifulSoup(html, 'html.parser')
    videoInfos = parseHTML.find_all(name='div', attrs={"class":"info"})

    for singleInfo in videoInfos:

        ## Get info from searching page

        # Div: img-anchor
        prv_sibling = singleInfo.previous_sibling
        IA = prv_sibling.find(name='span', attrs={"class":"so-imgTag_rb"})
        videoLen = IA.get_text()#In array

        # Div: headline clearfix
        HC = singleInfo.find(name='div', attrs={"class":"headline clearfix"})
        #subclass under headline clearfix
        videoType = HC.find(name='span', attrs={"class":"type hide"}).get_text()#In array
        titleMother = HC.find(name='a', attrs={"class":"title"})
        title = titleMother['title']#In array
        videoLink = titleMother['href']
        videoBVid = getBVNum(videoLink)#In array
        videoLink = getFullLink(videoLink)#In array

        # Div: des hide
        DH = singleInfo.find(name='div', attrs={"class":"des hide"})
        subscription = DH.get_text()#In array

        # Div: tags
        T = singleInfo.find(name='div', attrs={"class":"tags"})
        uploadTime = T.find(name='span', attrs={"class":"so-icon time"}).get_text()#In array
        Upinfos = T.find(name='a', attrs={"class":"up-name"})
        UpName = Upinfos.get_text()#In array
        UpLink = Upinfos['href']
        UpUUID = getUpUUID(UpLink)#In array
        UpLink = getFullLink(UpLink)#In array

        ## Get info of API-bilibili
        api_info = get_videoInfo(videoBVid)
        api_tagInfo, api_tag = get_tagInfo(videoBVid)
        api_upInfo = get_up_info(UpUUID)

        ## Put infos into a dictionary
        # Video info
        infoDict = {}
        infoDict['视频分类'] = videoType
        infoDict['视频时常'] = videoLen
        infoDict['视频标题'] = title
        infoDict['视频url'] = videoLink
        infoDict['视频BV号'] = videoBVid
        infoDict['视频AV号'] = api_info.get("aid")
        infoDict['视频简介'] = subscription
        infoDict['视频上传时间'] = uploadTime
        if (api_info.get("no_reprint") == 0):
            infoDict['视频可转载'] = "不可转载"#作品可转载？
        else: infoDict['视频可转载'] = "可转载"
        if (api_info.get("copyright") == 1):
            infoDict['视频类型'] = "自制"#作品类型：自制/转载
        else:   infoDict['视频类型'] = "转载"
        infoDict['视频标签'] = api_tag
        infoDict['视频标签详细信息'] = api_tagInfo
        # Viewer info
        infoDict['播放数'] = api_info.get("view")
        infoDict['弹幕数'] = api_info.get("danmaku")
        infoDict['点赞数'] = api_info.get("like")
        infoDict['投币数'] = api_info.get("coin")
        infoDict['收藏数'] = api_info.get("favorite")#收藏
        infoDict["分享数"] = api_info.get("share")
        infoDict['评论数'] = api_info.get("reply")
        infoDict["本周视频排行"] = api_info.get("now_rank")
        infoDict["历史视频排行"] = api_info.get("his_rank")
        # Up info
        infoDict['Up名字'] = UpName
        infoDict['Up粉丝数'] = api_upInfo.get("follower")#粉丝数
        infoDict['Up关注列表数'] = api_upInfo.get("following")#Up关注数
        infoDict['Up空间URL'] = UpLink
        infoDict['Up的uuid'] = UpUUID

        for key, value in infoDict.items():
            value = str(value)
            value = value.replace('\n', '')
            infoDict[key] = value.strip()

        # print(infoDict)
        videoList.append(infoDict)

def startSearch(keywords = '-海遥-'):
    
    global browser

    url = 'https://www.bilibili.com/'
    browser.get(url)
    # 获取输入框的id，并输入关键字python爬虫
    browser.find_element_by_class_name('nav-search-keyword').send_keys(keywords)
    # 输入回车进行搜索
    browser.find_element_by_class_name('nav-search-keyword').send_keys(Keys.ENTER)

    #切换为当前页，并关闭b站首页
    browser.close()
    handles = browser.window_handles
    browser.switch_to.window(handles[-1])

def selectMethod(dataType = 1):
    global videoList
    videoList = []
    # Type can be: 1.最新发布 2.最多播放

    #点选按最新发布排序
    if (dataType == 1):
        # 显示等待排序元素加载完成
        while True:
            try:
                buttNew = browser.find_element_by_xpath("//ul[@class='filter-type clearfix order']//li[3]")
                buttNew.click()
                time.sleep(1)
                break
            except: continue
    #点选按最多播放排序
    elif (dataType == 2):
        # 显示等待排序元素加载完成
        while True:
            try:
                buttNew = browser.find_element_by_xpath("//ul[@class='filter-type clearfix order']//li[2]") 
                buttNew.click()
                time.sleep(1)
                break
            except: continue

    #遍历所有页面，找到所有视频链接
    pageIndex = 1
    while True:
        
        try:
            # 显示等待排序元素加载完成
            while True:
                try:
                    fp_next = browser.find_element_by_xpath("//button[@class='nav-btn iconfont icon-arrowdown3']")
                    break
                except: pass

            print("Start searching at search-page", pageIndex, "/ 50", end='\r')

            #获取当前页面句柄
            normal_window = browser.current_window_handle
            #获取所有页面句柄
            all_Handles = browser.window_handles
            #如果新的pay_window句柄不是当前句柄，用switch_to_window方法切换
            for pay_window in all_Handles:
                if pay_window != normal_window:
                    browser.switch_to_window(pay_window)

            # 遍历所有视频，找到每个视频的详细信息
            # html = browser.execute_script("return document.documentElement.outerHTML")
            # print(html)
            getVideoLinks(browser.page_source)

            #Update page index
            pageIndex += 1

            # 找到下一页的元素page-item next
            fp_next = browser.find_element_by_xpath("//button[@class='nav-btn iconfont icon-arrowdown3']")
            # 点击下一页
            fp_next.click()

            # print("Finish searching at search-page", pageIndex, "/ 50")

            if (pageIndex == 50):
                break

        except:
            continue

## End section

# Section: writing information into excel

def writingInfo(sheet):

    global videoList
    # Initialize excel
    index = 1
    TitleRow = [i for i in videoList[0].keys()]
    for i in range(len(TitleRow)):
        sheet.cell(row = index, column = i+1).value = TitleRow[i]
    index += 1

    for subDict in videoList:
        colIndex = 1
        for subKey, subValue in subDict.items():
            try:
                sheet.cell(row = index, column = colIndex).value = subValue
            except:
                try:
                    print("row:", index, "column:", colIndex, "content:",subValue)
                except:
                    print("Invalid Input at", "row:", index, "column:", colIndex)
            colIndex += 1
        print("Writing data at", index, "/", len(videoList), end='\r')
        index += 1

def getWorkbook(dataType = 1):
    global book

    if (dataType == 1):
        sheet1 = book.create_sheet("MostNew", 0)
        writingInfo(sheet1)
    elif (dataType == 2):
        sheet2 = book.create_sheet("MostView", 1)
        writingInfo(sheet2)

# End section

# Section: main process

browser = webdriver.Firefox(executable_path = '/usr/local/bin/geckodriver')
videoList = []

if __name__ == "__main__":
    print("Program is started")

    keyword = '大司马'#put yout searching keyword here

    book = Workbook()
    today = time.strftime("%Y-%m-%d", time.localtime())
    bookname = "Video_about "+keyword+" at "+today+".xlsx"
    book.save(bookname)

    startSearch(keyword)
    
    # # Search result by most new
    # print('With sorting method: most new')
    # print("Start searching and finding all video ulr:")
    # selectMethod(1)
    # print("Start finding detailed info of each video:")
    # getWorkbook(1)
    # book.save(bookname)
    # print("Most new data is finished writing")
    
    # Search result by most view
    print('With sorting method: most view')
    print("Start searching and finding all video ulr:")
    selectMethod(2)
    print("Start finding detailed info of each video:")
    getWorkbook(2)
    book.save(bookname)
    print("Most view data is finished writing")

    book.close()
    browser.close()
    print("Program finished")