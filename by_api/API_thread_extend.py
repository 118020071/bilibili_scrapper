import requests, random, selenium, bs4, time, openpyxl, datetime, threading, concurrent.futures
from openpyxl import Workbook

# Section: get detailed info via API

g_headers = [
    "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
    'Opera/9.25 (Windows NT 5.1; U; en)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
    'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9',
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "
]

def get_json(bvid, url):

    global g_headers

    url = url+bvid

    fakeua = { "User-Agent":g_headers[random.randint(0, len(g_headers)-1)] }

    html = requests.get(
        url=url,
        headers=fakeua,
        timeout = 60
    )

    data = html.json()
    html.close()
    #return dict{}
    return data

def get_aid(bvid):

    url = "https://api.bilibili.com/x/web-interface/view?bvid="
    
    data = get_json(bvid, url)
    aid = data.get("data").get("aid")

    return aid

def get_videoInfo(bvid, video_data):
    
    url = "https://api.bilibili.com/x/web-interface/archive/stat?bvid="

    data = get_json(bvid, url)
    
    # form: aid,bvid,view,danmaku,reply,favorite,coin,share,like,now_rank,
    # his_rank,no_reprint,copyright,argue_msg,evaluation
    video_data = data.get("data")
    #copyright == 1: 自制, 2: 转载

    # print(video_data)
    # return video_data

def get_tagInfo(bvid, api_tagInfo, api_tag):
    
    url = "https://api.bilibili.com/x/web-interface/view/detail/tag?bvid="
    
    data = get_json(bvid, url)

    tag_data = data.get("data")

    whole_list = []
    tag_list = []

    for i in range(len(tag_data)):
        single_tag = tag_data[i]
        single_dict = {}
        single_dict["标签名称"] = single_tag.get("tag_name")
        single_dict["标签订阅数"] = single_tag.get("subscribed_count")
        single_dict["标签下视频数"] = single_tag.get("archive_count")
        single_dict["标签下精选视频数"] = single_tag.get("featured_count")

        whole_list.append(str(single_dict))
        tag_list.append(single_tag.get("tag_name"))

    # print(';'.join(whole_list), ','.join(tag_list))
    api_tagInfo = whole_list
    api_tag = tag_list
    # return ';'.join(whole_list), ','.join(tag_list)

def get_up_info(uuid, up_info):

    url = "https://api.bilibili.com/x/relation/stat?vmid="

    data = get_json(uuid, url)

    up_data = data.get("data")

    # up_info = {}
    # up_info[0] = up_data.get("following")
    # up_info[1] = up_data.get("follower")

    up_info["following"] = up_data.get("following")
    up_info["follower"] = up_data.get("follower")

    # print(up_info)
    # return up_info

# End section

# Section: reform data

def deleteUnit(a):
    #例如有'万',则添加0
    if "万" in a:
        aTenThousand = a[0:a.find(".")]
        aLeftover = a[(a.find(".")+1):a.find("万")]
        a = int(aTenThousand)*10000 + int(aLeftover)*1000
    return str(a)

def getBVNum(a):
    #get BV号 from link url
    a = a.split("/")
    a = a[-1].split("?")
    a = a[0]

    return a

def getFullLink(a):
    #get full link from link url
    a = a.split("?")
    a = a[0]
    a = "https:"+a
    return a

def getUpUUID(a):
    #input up link
    back = a.split("/")[-1]
    font = back.split("?")[0]

    return font

def unixTimeDecoding(value):
    #convert time in unix to yy:mm:dd hh:mm:ss
    format = '%Y-%m-%d %H:%M:%S'
    value = time.localtime(value)
    dt = time.strftime(format, value)

    return dt

# End section

# Section: get video BV id list using selenium

def getVideoDetails(pageNum, result, i):
    global videoList

    # //mode singleInfo.get("")
    singleInfo = result[i]
    ## Get info from searching page

    # Div: img-anchor
    videoLen = singleInfo.get("duration")#In array

    # Div: headline clearfix
    #subclass under headline clearfix
    videoType = singleInfo.get("typename")#In array
    title = singleInfo.get("title")#In array
    title = title.replace('<em class="keyword">',"")
    title = title.replace('</em>',"")
    videoBVid = singleInfo.get("bvid")
    videoLink = singleInfo.get("arcurl")
    if ("https" not in videoLink) and ("http" in videoLink):
        videoLink = videoLink.replace("http", "https")#In array

    # Div: des hide
    subscription = singleInfo.get("description")#In array

    # Div: tags
    uploadTime = singleInfo.get("pubdate")
    uploadTime = unixTimeDecoding(int(uploadTime))#In array
    updateTime = singleInfo.get("senddate")
    updateTime = unixTimeDecoding(int(updateTime))#In array
    
    UpName = singleInfo.get("author")#In array
    UpUUID = singleInfo.get("mid")#In array
    UpLink = "https://space.bilibili.com/"+str(UpUUID)#In array

    ## Get info of API-bilibili
    # api_info = {}
    # api_info_thread = threading.Thread(target=get_videoInfo, 
    # args=(videoBVid, api_info))
    # get_videoInfo(videoBVid,api_info)

    # api_tagInfo = []
    # api_tag = []
    # api_tag_thread = threading.Thread(target=get_tagInfo, 
    # args=(videoBVid, api_tagInfo, api_tag))
    # api_tagInfo = ';'.join(api_tagInfo)
    # api_tag = ','.join(api_tag)
    # api_tagInfo, api_tag = get_tagInfo(videoBVid)

    # api_upInfo = [0, 0]
    # api_upInfo_thread = threading.Thread(target=get_up_info,
    # args=(UpUUID, api_upInfo))
    # get_up_info(str(UpUUID), api_upInfo)
    # api_upInfo = get_up_info(UpUUID)

    ## Put infos into a dictionary
    # Video info
    infoDict = {}
    infoDict['搜索页位置'] = pageNum
    infoDict['UnixTimeStamp'] = singleInfo.get("pubdate")
    infoDict['视频分类'] = videoType
    infoDict['视频时常'] = videoLen
    infoDict['视频标题'] = title
    infoDict['视频url'] = videoLink
    infoDict['视频BV号'] = videoBVid
    infoDict['视频AV号'] = singleInfo.get("aid")
    infoDict['视频简介'] = subscription
    infoDict['视频上传时间'] = uploadTime
    infoDict['视频更新时间'] = updateTime

    #Sync thread
    # api_info_thread.join()
    # api_upInfo_thread.join()

    # if (api_info.get("no_reprint") == 0):
    #     infoDict['视频可转载'] = "不可转载"#作品可转载？####Sync
    # else: infoDict['视频可转载'] = "可转载"####Sync
    # if (api_info.get("copyright") == 1):
    #     infoDict['视频类型'] = "自制"#作品类型：自制/转载####Sync
    # else:   infoDict['视频类型'] = "转载"####Sync

    infoDict['视频标签'] = singleInfo.get("tag")
    # infoDict['视频标签详细信息'] = api_tagInfo
    # Viewer info
    infoDict['播放数'] = singleInfo.get("play")
    infoDict['弹幕数'] = singleInfo.get("video_review")

    # infoDict['点赞数'] = api_info.get("like")####Sync
    # infoDict['投币数'] = api_info.get("coin")####Sync

    infoDict['收藏数'] = singleInfo.get("favorites")#收藏

    # infoDict["分享数"] = api_info.get("share")####Sync

    infoDict['评论数'] = singleInfo.get("review")

    # infoDict["本周视频排行"] = api_info.get("now_rank")
    # infoDict["历史视频排行"] = api_info.get("his_rank")

    # Up info
    infoDict['Up名字'] = UpName
    if singleInfo.get("is_union_video") == 0:
        infoDict['是否联合投稿'] = "否"
    else:
        infoDict['是否联合投稿'] = "是"
    infoDict['视频总体排位分数'] = singleInfo.get("rank_score")

    # infoDict['Up粉丝数'] = api_upInfo[0]#粉丝数####Sync
    # infoDict['Up关注列表数'] = api_upInfo[1]#Up关注数####Sync

    infoDict['Up空间URL'] = UpLink
    infoDict['Up的uuid'] = UpUUID

    for key, value in infoDict.items():
        value = str(value)
        value = value.replace('\n', '')
        infoDict[key] = value.strip()

    # print(infoDict)
    videoList.append(infoDict)

# def getVideoDetails_api(i):
#     global videoList

def getVideoLinks(html):
    global threadList

    pageNum = str(html.get("page"))
    pageSize = int(html.get("pagesize"))
    result = html.get("result") #list object, with dict element

    for i in range(0, pageSize):
        x = threading.Thread(target=getVideoDetails,  args=(pageNum, result, i))
        x.start()
        threadList.append(x)
        
def startSearch(keywords = '-海遥-', searchType = 'click', duration = 0):

    global threadList
    
    url = "https://api.bilibili.com/x/web-interface/search/type?;order="+searchType+"&search_type=video&keyword="+keywords+"&duration="+str(duration)+"&page="

    for i in range(1,51):
        print("Start searching at page %i/50"%i, end='\r')
        data = get_json(str(i), url)
        search_result = data.get("data")
        getVideoLinks(search_result)

    for i in threadList:
        i.join()


## End section

# Section: writing information into excel

def writingInfo(sheet):

    global videoList
    # Initialize excel
    index = 1
    TitleRow = [i for i in videoList[0].keys()]
    for i in range(len(TitleRow)):
        sheet.cell(row = index, column = i+1).value = TitleRow[i]
    index += 1

    for subDict in videoList:
        colIndex = 1
        for subKey, subValue in subDict.items():
            try:
                sheet.cell(row = index, column = colIndex).value = subValue
            except:
                try:
                    print("row:", index, "column:", colIndex, "content:",subValue)
                except:
                    print("Invalid Input at", "row:", index, "column:", colIndex)
            colIndex += 1
        print("Writing data at", index, "/", len(videoList), end='\r')
        index += 1

def getWorkbook(dataType = 1):
    global book, videoList

    if (dataType == 1):
        videoList = sorted(videoList, key=lambda k: k['UnixTimeStamp'], reverse=True) 
        sheet1 = book.create_sheet("MostNew", 0)
        writingInfo(sheet1)
    elif (dataType == 2):
        sheet2 = book.create_sheet("MostView", 1)
        writingInfo(sheet2)

# End section

# Section: main process

videoList = []
threadList = []

if __name__ == "__main__":
    print("Program is started")

    keyword = '周淑怡'#put yout searching keyword here

    book = Workbook()
    today = time.strftime("%Y-%m-%d", time.localtime())
    bookname = "Video_about "+keyword+" at "+today+".xlsx"
    book.save(bookname)
    
    # # Search result by most new
    print('With sorting method: most new')
    print("Start searching and finding all video ulr:")
    startSearch(keyword, 'pubdate', 0)
    print("Start finding detailed info of each video:")
    getWorkbook(1)
    book.save(bookname)
    print("Most new data is finished writing")
    
    # Search result by most view
    # print('With sorting method: most view')
    # print("Start searching and finding all video ulr:")
    # startSearch(keyword, 'click', 0)
    # print("Start finding detailed info of each video:")
    # getWorkbook(2)
    # book.save(bookname)
    # print("Most view data is finished writing")

    # book.close()
    # print("Program finished")